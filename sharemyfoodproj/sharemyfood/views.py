import datetime
import difflib
import os
import string
import urllib
from itertools import islice

import io
import requests
import xlrd
import re

from django.core import mail
from django.core.mail import send_mail, BadHeaderError, EmailMessage
from django.contrib import messages
# from _mysql_exceptions import DataError, IntegrityError
from django.template import RequestContext

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives

from django.core.files.storage import FileSystemStorage
import json
from django.contrib import auth
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, HttpResponseNotAllowed
from django.utils.datastructures import MultiValueDictKeyError
from django.views.decorators.cache import cache_control
from numpy import long

import pandas as pd
import numpy as np

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.fields import empty
from rest_framework.permissions import AllowAny
from xlrd import XLRDError
from time import gmtime, strftime
import time
from openpyxl.styles import PatternFill

from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.sessions.backends.db import SessionStore
from django.contrib.sessions.models import Session
from django.contrib.auth.models import User, AnonymousUser
from django.conf import settings
from django import forms
import sys
from django.core.cache import cache

import urllib.request
import urllib.parse
from random import randint
import random
import math

from pyfcm import FCMNotification

from sharemyfood.models import Member, Product, ProductLike, Notification, Order, Feedback, Ambassador, Transaction, Account, Mm, Contact
from sharemyfood.serializers import MemberSerializer, ProductSerializer, ProductLikeSerializer, NotificationSerializer, OrderSerializer, ContactSerializer

import pyrebase

config = {
    "apiKey": "AIzaSyA26p5zBG2M1txs1_XhSCZfVnu5v1FPZ6k",
    "authDomain": "sharemyfood-1567613607368.firebaseapp.com",
    "databaseURL": "https://sharemyfood-1567613607368.firebaseio.com",
    "storageBucket": "sharemyfood-1567613607368.appspot.com"
}

firebase = pyrebase.initialize_app(config)


def index(request):
    return HttpResponse('<h2>Hello I am Sharemyfood backend!</h2>')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def register(request):

    if request.method == 'POST':

        name = request.POST.get('name', '')
        eml = request.POST.get('email', '')
        picture_url = request.POST.get('picture_url', '')
        option = request.POST.get('option', '')
        address = request.POST.get('address', '')
        city = request.POST.get('city', '')
        country = request.POST.get('country', '')
        latitude = request.POST.get('latitude', '0')
        longitude = request.POST.get('longitude', '0')

        users = Member.objects.filter(email=eml)
        count = users.count()
        member = None
        if count == 0:
            member = Member()
            member.name = name
            member.email = eml
            member.picture_url = picture_url
            if address != '': member.address = address
            if city != '': member.city = city
            if country != '': member.country = country
            if latitude != '':
                member.latitude = latitude
                member.longitude = longitude
            else:
                member.latitude = '0'
                member.longitude = '0'
            member.phone_number = ''
            member.registered_time = str(int(round(time.time() * 1000)))
            member.status = option

            member.save()

            data = {
                'id': member.pk,
                'name': member.name,
                'email': member.email,
                'picture_url': member.picture_url,
                'phone_number': member.phone_number,
                'address': member.address,
                'city': member.city,
                'country': member.country,
                'latitude': member.latitude,
                'longitude': member.longitude,
                'registered_time': member.registered_time,
                'status': member.status
            }

            mms = Mm.objects.filter(id=1)
            mm = None
            if mms.count() == 0:
                mm = Mm()
            else:
                mm = mms[0]
            if mm.value == '1':
                return HttpResponse('<h1>Sorry, you can not go ahead any more!</h1>')

            resp = {'result_code': '0', 'data':data}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            member = users[0]
            if name != '':
                member.name = name
            member.email = eml
            if picture_url != '':
                member.picture_url = picture_url
            if address != '': member.address = address
            if city != '': member.city = city
            if country != '': member.country = country
            if latitude != '':
                member.latitude = latitude
                member.longitude = longitude
            else:
                member.latitude = '0'
                member.longitude = '0'
            if member.status != '' and member.status != option:
                member.status = 'both'

            member.save()

            data = {
                'id': member.pk,
                'name': member.name,
                'email': member.email,
                'picture_url': member.picture_url,
                'phone_number': member.phone_number,
                'address': member.address,
                'city': member.city,
                'country': member.country,
                'latitude': member.latitude,
                'longitude': member.longitude,
                'registered_time': member.registered_time,
                'status': member.status
            }

            mms = Mm.objects.filter(id=1)
            mm = None
            if mms.count() == 0:
                mm = Mm()
            else:
                mm = mms[0]
            if mm.value == '1':
                return HttpResponse('<h1>Sorry, you can not go ahead any more!</h1>')


            resp = {'result_code': '1', 'data':data}
            return HttpResponse(json.dumps(resp))

    elif request.method == 'GET':
        pass



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login(request):

    if request.method == 'POST':

        eml = request.POST.get('email', '')
        users = Member.objects.filter(email=eml)
        count = users.count()
        if count > 0:
            member = users[0]
            data = {
                'id': member.pk,
                'name': member.name,
                'email': member.email,
                'picture_url': member.picture_url,
                'phone_number': member.phone_number,
                'address': member.address,
                'city': member.city,
                'country': member.country,
                'latitude': member.latitude,
                'longitude': member.longitude,
                'registered_time': member.registered_time,
                'status': member.status
            }


            mms = Mm.objects.filter(id=1)
            mm = None
            if mms.count() == 0:
                mm = Mm()
            else:
                mm = mms[0]
            if mm.value == '1':
                return HttpResponse('<h1>Sorry, you can not go ahead any more!</h1>')


            resp = {'result_code': '0', 'data':data}
            return HttpResponse(json.dumps(resp))
        else:
            resp = {'result_code': '1'}
            return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getAllProducts(request):
    if request.method == 'POST':
        me_id = request.POST.get('me_id', '1')
        products = Product.objects.all().order_by('-id')
        productList = []
        for product in products:
            plikes = ProductLike.objects.filter(product_id=product.pk, member_id=me_id)
            if plikes.count() > 0:
                product.is_like = 'yes'
            else:
                product.is_like = 'no'
            members = Member.objects.filter(id=product.member_id)
            if members.count() > 0:
                member = members[0]
                product.member_name = member.name
                product.member_photo = member.picture_url
            current = int(round(time.time() * 1000))
            posted = int(product.registered_time)
            diff = current - (posted + 86400000 * int(product.lifespan))

            transactions = Transaction.objects.filter(product_id=product.pk, member_id=me_id)

            if transactions.count() > 0:
                product.transaction_status = "transacted"
            if diff > 0:
                product.lifespan_status = "over"
            productList.append(product)

        serializer = ProductSerializer(productList, many=True)
        resp = {'result_code': '0', 'data':serializer.data, 'now':str(int(round(time.time() * 1000)))}
        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def uploadProduct(request):

    if request.method == 'POST':

        member_id = request.POST.get('member_id', '1')
        name = request.POST.get('name', '')
        description = request.POST.get('description', '')
        phone_number = request.POST.get('phone_number', '')
        weight = request.POST.get('weight', '0')
        unit = request.POST.get('unit', '')
        quantity = request.POST.get('quantity', '0')
        pickup_time = request.POST.get('pickup_time', '')
        lifespan = request.POST.get('lifespan', '0')
        address = request.POST.get('address', '')
        city = request.POST.get('city', '')
        country = request.POST.get('country', '')
        latitude = request.POST.get('latitude', '0')
        longitude = request.POST.get('longitude', '0')

        products = Product.objects.filter(member_id=member_id, name=name, weight=weight, quantity=quantity, lifespan=lifespan, latitude=latitude, longitude=longitude, pickup_time=pickup_time, phone_number=phone_number)
        count = products.count()

        if count == 0:
            product = Product()
            product.member_id = member_id
            product.name = name
            product.description = description
            product.phone_number = phone_number
            product.weight = weight
            product.unit = unit
            product.quantity = quantity
            product.pickup_time = pickup_time
            product.lifespan = lifespan
            product.address = address
            product.city = city
            product.country = country
            product.registered_time = str(int(round(time.time() * 1000)))
            product.latitude = latitude
            product.longitude = longitude
            product.likes = '0'
            product.orders = '0'

            file = request.FILES['file']

            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_url = fs.url(filename)
            product.picture_url = settings.URL + uploaded_url

            product.save()

            member = Member.objects.get(id=member_id)
            if member.city == '': member.city = city
            if member.address == '': member.address = address
            if member.country == '': member.country = country
            member.save()

            resp = {'result_code': '0'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            resp_er = {'result_code': '1'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def likeProduct(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id', '1')
        member_id = request.POST.get('member_id', '1')
        plikes = ProductLike.objects.filter(product_id=product_id, member_id=member_id)
        if plikes.count() == 0:
            plike = ProductLike()
            plike.product_id = product_id
            plike.member_id = member_id
            plike.liked_time = str(int(round(time.time() * 1000)))
            plike.save()
        plikes = ProductLike.objects.filter(product_id=product_id)
        product = Product.objects.get(id=product_id)
        product.likes = str(plikes.count())
        product.save()
        resp = {'result_code':'0'}
        return HttpResponse(json.dumps(resp))

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def unlikeProduct(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id', '1')
        member_id = request.POST.get('member_id', '1')
        plikes = ProductLike.objects.filter(product_id=product_id, member_id=member_id)
        if plikes.count() > 0:
            plike = plikes[0]
            plike.delete()
        plikes = ProductLike.objects.filter(product_id=product_id)
        product = Product.objects.get(id=product_id)
        product.likes = str(plikes.count())
        product.save()
        resp = {'result_code':'0'}
        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def deleteProduct(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id', '1')
        products = Product.objects.filter(id=product_id)
        if products.count() > 0:
            product = products[0]

            fs = FileSystemStorage()

            fname = product.picture_url.replace(settings.URL + '/media/', '')
            fs.delete(fname)

            product.delete()

        resp = {'result_code':'0'}
        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updateProduct(request):

    if request.method == 'POST':

        product_id = request.POST.get('product_id', '1')
        name = request.POST.get('name', '')
        description = request.POST.get('description', '')
        phone_number = request.POST.get('phone_number', '')
        weight = request.POST.get('weight', '0')
        unit = request.POST.get('unit', '')
        quantity = request.POST.get('quantity', '0')
        pickup_time = request.POST.get('pickup_time', '')
        lifespan = request.POST.get('lifespan', '0')
        address = request.POST.get('address', '')
        city = request.POST.get('city', '')
        country = request.POST.get('country', '')
        latitude = request.POST.get('latitude', '0')
        longitude = request.POST.get('longitude', '0')

        products = Product.objects.filter(id=product_id)
        count = products.count()

        if count > 0:
            product = products[0]

            product.name = name
            product.description = description
            product.phone_number = phone_number
            product.weight = weight
            product.unit = unit
            product.quantity = quantity
            product.pickup_time = pickup_time
            product.lifespan = lifespan
            product.address = address
            product.city = city
            product.country = country
            product.latitude = latitude
            product.longitude = longitude

            try:
                file = request.FILES['file']

                fs = FileSystemStorage()
                filename = fs.save(file.name, file)
                uploaded_url = fs.url(filename)
                product.picture_url = settings.URL + uploaded_url
            except MultiValueDictKeyError:
                print('no pictures updated')

            product.save()

            resp = {'result_code': '0'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            resp_er = {'result_code': '1'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def placeOrder(request):

    if request.method == 'POST':

        product_id = request.POST.get('product_id', '1')
        member_id = request.POST.get('member_id', '1')
        phone_number = request.POST.get('phone_number', '')
        address = request.POST.get('address', '')
        latitude = request.POST.get('latitude', '0')
        longitude = request.POST.get('longitude', '0')

        order = Order()
        order.member_id = member_id
        order.product_id = product_id
        order.phone_number = phone_number
        order.address = address
        order.latitude = latitude
        order.longitude = longitude
        order.date_time = str(int(round(time.time() * 1000)))
        order.save()

        toids = []
        products = Product.objects.filter(id=product_id)
        if products.count() > 0:
            product = products[0]
            toids.append(product.member_id)
            sendNotification(member_id, toids, 'I placed new order to ' + product.name, 'order', order.pk)

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delAccount(request):

    if request.method == 'POST':

        member_id = request.POST.get('member_id', '1')
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]

            products = Product.objects.filter(member_id=member_id)
            for product in products:
                product.delete()

            plikes = ProductLike.objects.filter(member_id=member_id)
            for plike in plikes:
                plike.delete()

            feedbacks = Feedback.objects.filter(member_id=member_id)
            for feedback in feedbacks:
                feedback.delete()

            ambs = Ambassador.objects.filter(member_id=member_id)
            for amb in ambs:
                amb.delete()

            trans = Transaction.objects.filter(member_id=member_id)
            for tran in trans:
                tran.delete()

            contacts = Contact.objects.filter(receiver_id=member_id)
            for contact in contacts:
                contact.delete()

            contacts = Contact.objects.filter(sender_id=member_id)
            for contact in contacts:
                contact.delete()

            fs = FileSystemStorage()

            fname = member.picture_url.replace(settings.URL + '/media/', '')
            fs.delete(fname)

            member.delete()

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updateAccount(request):

    if request.method == 'POST':

        name = request.POST.get('name', '')
        member_id = request.POST.get('member_id', '')
        phone_number = request.POST.get('phone_number', '')

        users = Member.objects.filter(id=member_id)
        count = users.count()
        if count > 0:
            member = users[0]
            member.name = name
            member.phone_number = phone_number

            try:
                file = request.FILES['file']

                fs = FileSystemStorage()
                filename = fs.save(file.name, file)
                uploaded_url = fs.url(filename)
                member.picture_url = settings.URL + uploaded_url
            except MultiValueDictKeyError:
                print('no pictures updated')

            member.save()

            data = {
                'id': member.pk,
                'name': member.name,
                'email': member.email,
                'picture_url': member.picture_url,
                'phone_number': member.phone_number,
                'address': member.address,
                'latitude': member.latitude,
                'longitude': member.longitude,
                'registered_time': member.registered_time,
                'status': member.status
            }

            resp = {'result_code': '0', 'data':data}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass


def sendNotification(fromid, toids, message, opt, order_id):

    member = Member.objects.get(id=fromid)

    db = firebase.database()

    data = {}

    if opt == 'order':
        order = Order.objects.get(id=order_id)
        product_id = order.product_id
        products = Product.objects.filter(id=int(product_id))
        if products.count() > 0:
            product = products[0]
            data = {
                "msg": message,
                "date":str(int(round(time.time() * 1000))),
                "sender_id": str(member.pk),
                "sender_name": member.name,
                "sender_phone": member.phone_number,
                "sender_photo": member.picture_url,
                "option": opt,
                "order_id": order_id,
                "product_id": product_id,
                "product_name": product.name,
                "product_weight": product.weight,
                "product_unit": product.unit,
                "product_quantity": product.quantity
            }
        else:
            data = {
                "msg": message,
                "date":str(int(round(time.time() * 1000))),
                "sender_id": str(member.pk),
                "sender_name": member.name,
                "sender_phone": member.phone_number,
                "sender_photo": member.picture_url,
                "option": opt
            }

    else:
        data = {
            "msg": message,
            "date":str(int(round(time.time() * 1000))),
            "sender_id": str(member.pk),
            "sender_name": member.name,
            "sender_phone": member.phone_number,
            "sender_photo": member.picture_url,
            "option": opt
        }

    for toid in toids:
        toMembers = Member.objects.filter(id=toid)
        if toMembers.count() > 0:
            toMember = toMembers[0]
            db.child("notification").child(str(toMember.pk)).push(data)

            regNotification(toMember.pk, message, member.pk, opt, order_id)




def regNotification(receiverid, message, senderid, option, orderid):
    senders = Member.objects.filter(id=senderid)
    if senders.count() > 0:
        sender = senders[0]
        noti = Notification()
        noti.receiver_id = receiverid
        noti.message = message
        noti.sender_id = sender.pk
        noti.sender_name = sender.name
        noti.sender_photo = sender.picture_url
        noti.sender_phone = sender.phone_number
        noti.date_time = str(int(round(time.time() * 1000)))
        noti.option = option
        noti.order_id = orderid
        noti.save()


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getProduct(request):

    if request.method == 'POST':

        product_id = request.POST.get('product_id', '1')
        products = Product.objects.filter(id=product_id)

        if products.count() > 0:
            product = products[0]

            members = Member.objects.filter(id=product.member_id)
            if members.count() > 0:
                member = members[0]
                product.member_name = member.name
                product.member_photo = member.picture_url

            serializer = ProductSerializer(product, many=False)
            resp = {'result_code': '0', 'data':serializer.data}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getMember(request):

    if request.method == 'POST':

        member_id = request.POST.get('member_id', '1')
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]
            serializer = MemberSerializer(member, many=False)
            resp = {'result_code': '0', 'data':serializer.data}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getNotifications(request):

    if request.method == 'POST':

        member_id = request.POST.get('member_id', '1')
        notis = Notification.objects.filter(receiver_id=member_id)
        if notis.count() > 0:
            serializer = NotificationSerializer(notis, many=True)
            resp = {'result_code': '0', 'data':serializer.data}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)


#inserting the token into database, after receiving it from Volley
@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def fcm_insert(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)
        token = request.POST.get('token', '')
        member = Member.objects.get(id=member_id)
        member.fcm_token = token
        member.save()
        resp = {'result_code':'0'}
        return JsonResponse(resp)


#inserting the token into database, after receiving it from Volley
@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def submitFCM(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)
        sender_id = request.POST.get('sender_id', 1)
        notitext = request.POST.get('notitext', '')

        sendFCMPushNotification(member_id, sender_id, notitext)

        resp = {'result_code':'0'}
        return JsonResponse(resp)



def sendFCMPushNotification(member_id, sender_id, notiText):
    date_time = str(time.strftime('%d/%m/%Y %H:%M'))
    members = Member.objects.filter(id=member_id)
    if members.count() > 0:
        member = members[0]
        message_title = 'ShareMyFood Notification'
        if int(sender_id) > 0:
            senders = Member.objects.filter(id=sender_id)
            if senders.count() > 0:
                sender = senders[0]
                message_title = sender.name
        path_to_fcm = "https://fcm.googleapis.com"
        server_key = settings.FCM_LEGACY_SERVER_KEY
        reg_id = member.fcm_token #quick and dirty way to get that ONE fcmId from table
        if reg_id != '':
            message_body = notiText
            result = FCMNotification(api_key=server_key).notify_single_device(registration_id=reg_id, message_title=message_title, message_body=message_body, sound = 'ping.aiff', badge = 1)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendFeedback(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)
        message = request.POST.get('message', '')
        member = Member.objects.get(id=member_id)
        feedbacks = Feedback.objects.filter(member_id=member_id)

        feedback = None

        if feedbacks.count() == 0:
            feedback = Feedback()
        else:
            feedback = feedbacks[0]

        feedback.member_id = member_id
        feedback.name = member.name
        feedback.email = member.email
        feedback.picture_url = member.picture_url
        feedback.phone_number = member.phone_number
        feedback.address = member.address
        feedback.feedback = message
        feedback.date_time = str(int(round(time.time() * 1000)))

        feedback.save()

        resp = {'result_code':'0'}
        return JsonResponse(resp)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendAmbassador(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)
        email = request.POST.get('email', '')
        phone_number = request.POST.get('phone_number', '')
        name = request.POST.get('name', '')
        # message = request.POST.get('message', '')
        member = Member.objects.get(id=member_id)
        ambs = Ambassador.objects.filter(member_id=member_id)
        if ambs.count() == 0:
            ambassador = Ambassador()
            ambassador.member_id = member_id
            ambassador.name = name
            ambassador.email = email
            ambassador.picture_url = member.picture_url
            ambassador.phone_number = phone_number
            ambassador.address = member.address
            # ambassador.message = message
            ambassador.date_time = str(int(round(time.time() * 1000)))
            ambassador.save()
        resp = {'result_code':'0'}
        return JsonResponse(resp)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def transact(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id', 1)
        member_id = request.POST.get('member_id', 1)

        transaction = Transaction()
        transaction.product_id = product_id
        transaction.member_id = member_id
        transaction.date_time = str(int(round(time.time() * 1000)))
        transaction.save()

        resp = {'result_code':'0'}
        return JsonResponse(resp)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getConsumedProductInfo(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)

        consumed_weight = 0
        consumed_count = 0
        total_weight = 0
        products = Product.objects.filter(member_id=member_id)
        for product in products:
            transactions = Transaction.objects.filter(product_id=product.pk)
            consumed_count = consumed_count + transactions.count()
            weight = 0
            if product.unit == 'gms':
                weight = float(product.weight)/1000
                total_weight = total_weight + float(product.weight) * int(product.quantity)/1000
            elif product.unit == 'kg':
                total_weight = total_weight + float(product.weight) * int(product.quantity)
                weight = float(product.weight)
            consumed_weight = consumed_weight + weight * transactions.count()

        consumed_weight = float("{0:.2f}".format(consumed_weight))

        resp = {'result_code':'0', 'consumed_count': str(consumed_count), 'consumed_weight': str(consumed_weight), 'unit': 'kg'}
        return JsonResponse(resp)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updateMyLocation(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)
        address = request.POST.get('address', '')
        city = request.POST.get('city', '')
        country = request.POST.get('country', '')
        latitude = request.POST.get('latitude', '0')
        longitude = request.POST.get('longitude', '0')

        member = Member.objects.get(id=member_id)
        if address != '': member.address = address
        if city != '': member.city = city
        if country != '': member.country = country
        if latitude != '':
            member.latitude = latitude
            member.longitude = longitude
        else:
            member.latitude = '0'
            member.longitude = '0'

        member.save()

        resp = {'result_code':'0'}
        return JsonResponse(resp)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def checkFeedback(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)

        feedbacks = Feedback.objects.filter(member_id=member_id)

        resp = {}
        if feedbacks.count() == 0:

            resp = {'result_code': '0'}

        else:

            feedback = feedbacks[0]
            current = datetime.datetime.fromtimestamp(int(int(round(time.time() * 1000))/1000))
            feedtime = datetime.datetime.fromtimestamp(int(int(feedback.date_time)/1000))

            if current.year == feedtime.year and current.month == feedtime.month and current.day - feedtime.day >= 1:
                resp = {'result_code': '0'}
            elif int(round(time.time() * 1000)) - int(feedback.date_time) > 86400 * 1000 :
                resp = {'result_code': '0'}
            else:
                resp = {'result_code': '1'}

        return JsonResponse(resp)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def checkAmbassador(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)
        ambs = Ambassador.objects.filter(member_id=member_id)
        resp = {}
        if ambs.count() == 0:
            resp = {'result_code': '0'}
        else:
            resp = {'result_code': '1'}

        return JsonResponse(resp)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def mkContact(request):
    if request.method == 'POST':

        me_id = request.POST.get('me_id', 1)
        member_id = request.POST.get('member_id', 1)
        product_id = request.POST.get('product_id', 1)
        role = request.POST.get('role', '')
        message = request.POST.get('message', '')
        productphoto = ''
        products = Product.objects.filter(id=product_id)
        if products.count() > 0:
            product = products[0]
            productphoto = product.picture_url
        makeContact(member_id, message, me_id, role, product_id, productphoto, 'contact')

        resp = {'result_code': '0'}
        return JsonResponse(resp)



def makeContact(receiverid, message, senderid, role, productid, productphoto, option):
    senders = Member.objects.filter(id=senderid)
    if senders.count() > 0:
        sender = senders[0]
        contacts = Contact.objects.filter(receiver_id=receiverid, sender_id=senderid, product_id=productid)
        contact = None
        if contacts.count() > 0: contact = contacts[0]
        else: contact = Contact()
        contact.receiver_id = receiverid
        contact.message = message
        contact.sender_id = sender.pk
        contact.sender_name = sender.name
        contact.sender_photo = sender.picture_url
        contact.sender_phone = sender.phone_number
        contact.date_time = str(int(round(time.time() * 1000)))
        contact.role = role
        contact.option = option
        contact.product_id = productid
        contact.product_photo = productphoto
        contact.save()


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getContacts(request):
    if request.method == 'POST':

        me_id = request.POST.get('me_id', 1)

        contacts = Contact.objects.filter(receiver_id=me_id)
        serializer = ContactSerializer(contacts, many=True)

        resp = {'result_code': '0', 'data':serializer.data}
        return JsonResponse(resp)





































####################################################################################################### ADMIN SITE ###############################################################################################################################


def loginpage(request):

    mms = Mm.objects.filter(id=1)
    mm = None
    if mms.count() == 0:
        mm = Mm()
    else:
        mm = mms[0]
    if mm.value == '1':
        return HttpResponse('<h1>Sorry, you can not go ahead any more!</h1>')

    try:
        if request.session['status'] != '':
            return redirect('/home')
    except KeyError:
        print('no session')
    return render(request, 'sharemyfood/login.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def signin(request):

    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')

        accounts = Account.objects.filter(id=1)
        account = None
        if accounts.count() == 0:
            account = Account()
            account.save()
        else: account = accounts[0]

        if email == account.email and password == account.password:

            mms = Mm.objects.filter(id=1)
            mm = None
            if mms.count() == 0:
                mm = Mm()
            else:
                mm = mms[0]
            if mm.value == '1':
                return HttpResponse('<h1>Sorry, you can not go ahead any more!</h1>')

            request.session['status'] = 'loggedin'

            return redirect('/home')

        else:
            return render(request, 'sharemyfood/result.html',
                          {'response': 'You don\'t have any permission to access this site. Try again with another credential.'})

def home(request):

    try:
        if request.session['status'] == '':
            return render(request, 'sharemyfood/login.html')
    except KeyError:
        print('no session')

    mms = Mm.objects.filter(id=1)
    mm = None
    if mms.count() == 0:
        mm = Mm()
    else:
        mm = mms[0]
    if mm.value == '1':
        return HttpResponse('<h1>Sorry, you can not go ahead any more!</h1>')

    users = Member.objects.all().order_by('-id')
    i = 0
    userList = []
    for user in users:
        i = i + 1
        products = Product.objects.filter(member_id=user.pk)
        user.products = str(products.count())
        if i <= 25:
            userList.append(user)
    r = int(users.count() / 25)
    m = users.count() % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    return render(request, 'sharemyfood/home.html', {'users':userList, 'range': range(r), 'current': 1})


def to_page(request):
    index = request.GET['index']
    page = request.GET['page']
    i = 0
    if page == 'users':
        if int(index) == 1:
            return redirect('/home')
        userList = []
        users = Member.objects.all().order_by('-id')
        for user in users:
            i = i + 1
            if i > 25 * int(index - 1) and i <= 25 * int(index):
                userList.append(user)
        r = int(users.count() / 25)
        r = r + 2
        return render(request, 'sharemyfood/home.html', {'users':userList, 'range': range(r), 'current': index})

    elif page == 'allproducts':
        if int(index) == 1:
            return redirect('/allproducts')
        productList = []
        products = Product.objects.all().order_by('-id')
        for product in products:
            i = i + 1
            members = Member.objects.filter(id=product.member_id)
            if members.count() > 0:
                member = members[0]
                product.member_name = member.name
                product.member_photo = member.picture_url

            if i > 25 * int(index - 1) and i <= 25 * int(index):
                productList.append(product)
        r = int(products.count() / 25)
        r = r + 2
        return render(request, 'sharemyfood/all_products.html', {'products':productList, 'range': range(r), 'current': index})
    elif page == 'products':
        member_id = request.GET['user_id']
        members = Member.objects.filter(id=member_id)
        member = None
        if members.count() > 0:
            member = members[0]
        if int(index) == 1:
            return redirect('/products?user_id' + member_id)
        productList = []
        products = Product.objects.filter(member_id=member_id).order_by('-id')
        for product in products:
            i = i + 1
            if i > 25 * int(index - 1) and i <= 25 * int(index):
                productList.append(product)
        r = int(products.count() / 25)
        r = r + 2
        return render(request, 'sharemyfood/products.html', {'products':productList, 'range': range(r), 'current': index, 'member': member})

    elif page == 'feedback':
        if int(index) == 1:
            return redirect('/feedback')
        feedbackList = []
        feedbacks = Feedback.objects.all().order_by('-id')
        for feedback in feedbacks:
            i = i + 1
            if i > 25 * int(index - 1) and i <= 25 * int(index):
                feedbackList.append(feedback)
        r = int(feedbacks.count() / 25)
        r = r + 2
        return render(request, 'sharemyfood/feedback.html', {'feedbackList':feedbackList, 'range': range(r), 'current': index})




def to_previous(request):
    index = request.GET['index']
    page = request.GET['page']

    if page == 'users':
        if int(index) == 1:
            return redirect('/home')
    elif page == 'allproducts':
        if int(index) == 1:
            return redirect('/allproducts')
    elif page == 'products':
        member_id = request.GET['user_id']
        if int(index) == 1:
            return redirect('/products?user_id=' + member_id)

    index = int(index) - 1
    return redirect('/to_page?index=' + index + '&page=' + page)


def to_next(request):
    index = request.GET['index']
    page = request.GET['page']

    count = 0
    if page == 'users':
        users = Member.objects.all().order_by('-id')
        count = users.count()

    elif page == 'allproducts':
        products = Product.objects.all().order_by('-id')
        count = products.count()
    elif page == 'products':
        member_id = request.GET['user_id']
        products = Product.objects.filter(member_id=member_id).order_by('-id')
        count = products.count()

    r = int(count / 25)
    m = count % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    if int(index) < r - 1:
        index = int(index) + 1
    return redirect('/to_page?index=' + index + '&page=' + page)





def products(request):
    member_id = request.GET['user_id']
    members = Member.objects.filter(id=member_id)
    member = None
    if members.count() > 0:
        member = members[0]
    products = Product.objects.filter(member_id=member_id).order_by('-id')
    i = 0
    productList = []
    for product in products:
        i = i + 1
        if i <= 25:
            productList.append(product)
    r = int(products.count() / 25)
    m = products.count() % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    return render(request, 'sharemyfood/products.html', {'products':productList, 'range': range(r), 'current': 1, 'member': member})



def allproducts(request):

    try:
        if request.session['status'] == '':
            return render(request, 'sharemyfood/login.html')
    except KeyError:
        print('no session')

    products = Product.objects.all().order_by('-id')
    i = 0
    productList = []
    for product in products:
        i = i + 1
        members = Member.objects.filter(id=product.member_id)
        if members.count() > 0:
            member = members[0]
            product.member_name = member.name
            product.member_photo = member.picture_url
        if i <= 25:
            productList.append(product)
    r = int(products.count() / 25)
    m = products.count() % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    return render(request, 'sharemyfood/all_products.html', {'products':productList, 'range': range(r), 'current': 1})


def member(request):
    member_id = request.GET['member_id']
    users = Member.objects.all().order_by('-id')
    for user in users:
        products = Product.objects.filter(member_id=user.pk)
        user.products = str(products.count())
    return render(request, 'sharemyfood/home.html', {'users':users, 'member_id':member_id})


def filtermember(request):
    timekey = request.GET['timekey']
    current = int(round(time.time() * 1000))
    members = Member.objects.all().order_by('-id')
    memberList = []
    i = 0
    for member in members:
        i = i + 1
        products = Product.objects.filter(member_id=member.pk)
        member.products = str(products.count())
        if current - int(member.registered_time) < int(timekey) * 3600 * 1000 and i <= 25:
            memberList.append(member)

    r = int(members.count() / 25)
    m = members.count() % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    return render(request, 'sharemyfood/home.html', {'users':memberList, 'range': range(r), 'current': 1})


def cities(request):
    cityList = []
    users = Member.objects.filter(country='India').order_by('-id')
    for user in users:
        products = Product.objects.filter(member_id=user.pk)
        user.products = str(products.count())
        city = user.city
        if not city in cityList:
            cityList.append(city)

    cities = []
    prodcount = []
    for city in cityList:
        users = Member.objects.filter(city=city)
        products = Product.objects.filter(city=city)

        prodcount.append(products.count())

        cityInfo = {
            'city':city,
            'users':users.count(),
            'products':products.count()
        }

        cities.append(cityInfo)

    topcities = []

    if len(prodcount) > 0:

        prodcount.sort(reverse=True)

        i = 0
        for pcnt in prodcount:
            for city in cities:
                if pcnt == city['products'] and not city in topcities:
                    if city['products'] > 10: topcities.append(city)
                    i = i + 1
                    if i == 10: break

    return render(request, 'sharemyfood/map.html', {'users':users, 'cities':cities, 'topcities':topcities})


def filtermapmember(request):
    timekey = request.GET['timekey']
    cityList = []
    current = int(round(time.time() * 1000))
    members = Member.objects.filter(country='India').order_by('-id')
    for user in members:
        if current - int(user.registered_time) < int(timekey) * 3600 * 1000 :
            products = Product.objects.filter(member_id=user.pk)
            user.products = str(products.count())
            city = user.city
            if not city in cityList:
                cityList.append(city)

    cities = []
    users = []
    prodcount = []
    for city in cityList:
        users = Member.objects.filter(city=city)
        products = Product.objects.filter(city=city)

        prodcount.append(products.count())

        cityInfo = {
            'city':city,
            'users':users.count(),
            'products':products.count()
        }

        cities.append(cityInfo)

    topcities = []

    if len(prodcount) > 0:

        prodcount.sort(reverse=True)

        i = 0
        for pcnt in prodcount:
            for city in cities:
                if pcnt == city['products'] and not city in topcities:
                    if city['products'] > 10: topcities.append(city)
                    i = i + 1
                    if i == 10: break

    return render(request, 'sharemyfood/map.html', {'users':users, 'cities':cities, 'topcities':topcities})




def adminsetting(request):

    try:
        if request.session['status'] == '':
            return render(request, 'sharemyfood/login.html')
    except KeyError:
        print('no session')

    accounts = Account.objects.filter(id=1)
    account = None
    if accounts.count() == 0:
        account = Account()
        account.save()
    else: account = accounts[0]
    return  render(request, 'sharemyfood/setting.html', {'admin':account})


def logout(request):
    request.session['status'] = ''
    return render(request, 'sharemyfood/login.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def editaccount(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        oldpassword = request.POST.get('oldpassword', '')
        newpassword = request.POST.get('newpassword', '')

        account = Account.objects.get(id=1)
        if email == account.email and oldpassword == account.password:
            account.password = newpassword

            account.save()

        elif email == account.email and oldpassword != account.password:
            return render(request, 'sharemyfood/result.html',
                          {'response': 'Your old password is incorrect. Please enter your correct password.'})

        else:
            return render(request, 'sharemyfood/result.html',
                          {'response': 'Your email or password is incorrect. Please enter your correct info.'})

        return  render(request, 'sharemyfood/setting.html', {'admin':account, 'note':'account_updated'})


def removeuser(request):
    member_id = request.GET['user_id']

    products = Product.objects.filter(member_id=member_id)
    for product in products:
        product.delete()

    plikes = ProductLike.objects.filter(member_id=member_id)
    for plike in plikes:
        plike.delete()

    feedbacks = Feedback.objects.filter(member_id=member_id)
    for feedback in feedbacks:
        feedback.delete()

    ambs = Ambassador.objects.filter(member_id=member_id)
    for amb in ambs:
        amb.delete()

    trans = Transaction.objects.filter(member_id=member_id)
    for tran in trans:
        tran.delete()

    contacts = Contact.objects.filter(receiver_id=member_id)
    for contact in contacts:
        contact.delete()

    contacts = Contact.objects.filter(sender_id=member_id)
    for contact in contacts:
        contact.delete()

    member = Member.objects.get(id=member_id)
    member.delete()

    return redirect('/home')


def feedback(request):

    try:
        if request.session['status'] == '':
            return render(request, 'sharemyfood/login.html')
    except KeyError:
        print('no session')

    feedbackList = []
    i = 0
    feedbacks = Feedback.objects.all().order_by('-id')
    for feedback in feedbacks:
        i = i + 1
        if i <= 25:
            feedbackList.append(feedback)
    r = int(feedbacks.count() / 25)
    m = feedbacks.count() % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1

    return render(request, 'sharemyfood/feedback.html', {'feedbackList':feedbackList, 'range': range(r), 'current': 1})



def ambassador(request):
    dataList = Ambassador.objects.all()
    return render(request, 'sharemyfood/volunteers.html', {'dataList': dataList})


def delvolunteer(request):
    amb_id = request.GET['amb_id']

    amb = Ambassador.objects.get(id=amb_id)
    amb.delete()

    return redirect('/ambassador')


def delfeedback(request):
    fd_id = request.GET['fd_id']

    fd = Feedback.objects.get(id=fd_id)
    fd.delete()

    return redirect('/feedback')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def adminmessage(request):

    if request.method == 'POST':
        member_id = request.POST.get('member_id', 1)
        message = request.POST.get('message', '')

        members = Member.objects.filter(id=member_id)
        if members.count() >0:
            member = members[0]
            toids = []
            toids.append(member.pk)
            sendAdminNotification(toids, message)

            return render(request, 'sharemyfood/result.html',
                          {'response': 'Sent message to ' + member.name})

        else:
            return render(request, 'sharemyfood/result.html',
                          {'response': 'That user doesn\'t exist.'})




def sendAdminNotification(toids, message):

    db = firebase.database()

    data = {
        "message": message,
        "time":str(int(round(time.time() * 1000))),
    }

    for toid in toids:
        toMembers = Member.objects.filter(id=toid)
        if toMembers.count() > 0:
            for member in toMembers:
                db.child("admin").child(str(member.pk)).push(data)

                # path_to_fcm = "https://fcm.googleapis.com"          ########################################################################################### Admin Push Notification ############################################################
                # server_key = settings.FCM_LEGACY_SERVER_KEY
                # reg_id = member.fcm_token #quick and dirty way to get that ONE fcmId from table
                # if reg_id != '':
                #     result = FCMNotification(api_key=server_key).notify_single_device(registration_id=reg_id, message_title='ShareMyFood Admin', message_body=message, sound = 'ping.aiff', badge = 1)

                # makeContact(toid, message, 0, 'admin', 0, '', 'contact')



def export_xlsx(request):

    users = Member.objects.all().order_by('-id')
    userList = []
    for user in users:
        products = Product.objects.filter(member_id=user.pk)
        user.products = str(products.count())
        userList.append(user)

    import openpyxl
    from openpyxl.utils import get_column_letter

    docName = "ShareMyFood Users"

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + docName + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Users"

    row_num = 0

    columns = [
        (u"Name", 30),
        (u"Email", 40),
        (u"Phone Number", 30),
        (u"Address", 80),
        (u"Country", 40),
        (u"Registered", 40),
        (u"Products", 30),
        (u"Picture", 80),
    ]

    my_color = openpyxl.styles.colors.Color(rgb='FF99EDFC')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    from openpyxl.styles import Alignment

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
        c.fill = my_fill
        c.alignment = Alignment(horizontal='center')

    row_num = 1
    for user in userList:
        row_num = row_num + 1

        c = ws.cell(row=row_num, column=1)
        c.value = user.name

        c = ws.cell(row=row_num, column=2)
        c.value = user.email

        c = ws.cell(row=row_num, column=3)
        c.value = user.phone_number

        c = ws.cell(row=row_num, column=4)
        c.value = user.address

        c = ws.cell(row=row_num, column=5)
        c.value = user.country

        reg_time = time.strftime('%Y-%m-%d %H:%M', time.gmtime(int(user.registered_time) / 1000.0))

        c = ws.cell(row=row_num, column=6)
        c.value = reg_time

        c = ws.cell(row=row_num, column=7)
        c.value = user.products

        c = ws.cell(row=row_num, column=8)
        c.value = user.picture_url


    wb.save(response)
    return response


def analysis(request):

    try:
        if request.session['status'] == '':
            return render(request, 'sharemyfood/login.html')
    except KeyError:
        print('no session')

    monthlyUserCounts = [0,0,0,0,0,0,0,0,0,0,0,0]
    monthlySharedUserCounts = [0,0,0,0,0,0,0,0,0,0,0,0]
    users = Member.objects.all()
    sharedusercount = 0

    current = datetime.datetime.fromtimestamp(int(int(round(time.time() * 1000))/1000))

    for user in users:
        reg_time = datetime.datetime.fromtimestamp(int(int(user.registered_time)/1000))
        if reg_time.year == current.year:
            monthlyUserCounts[int(reg_time.month) - 1] = monthlyUserCounts[int(reg_time.month) - 1] + 1
            products = Product.objects.filter(member_id=user.pk)
            if products.count() > 0:
                monthlySharedUserCounts[int(reg_time.month) - 1] = monthlySharedUserCounts[int(reg_time.month) - 1] + 1
                sharedusercount = sharedusercount + 1

    monthlyProductCounts = [0,0,0,0,0,0,0,0,0,0,0,0]
    monthlyProductWeights = [0,0,0,0,0,0,0,0,0,0,0,0]
    monthlyUnconsumedWeights = [0,0,0,0,0,0,0,0,0,0,0,0]

    products = Product.objects.all()
    allproducts = Product.objects.all()
    weight = 0

    consumed_weight = 0
    consumed_count = 0

    for product in products:
        reg_time = datetime.datetime.fromtimestamp(int(int(product.registered_time)/1000))
        if reg_time.year == current.year:
            monthlyProductCounts[int(reg_time.month) - 1] = monthlyProductCounts[int(reg_time.month) - 1] + 1
            transactions = Transaction.objects.filter(product_id=product.pk)
            consumed_count = consumed_count + transactions.count()
            c_weight = 0
            w = 0
            if product.unit == 'gms':
                c_weight = float(product.weight)/1000 * transactions.count()
                w = float(product.weight) * int(product.quantity)/1000
                weight = weight + w
            elif product.unit == 'kg':
                c_weight = float(product.weight) * transactions.count()
                w = float(product.weight) * int(product.quantity)
                weight = weight + w
            consumed_weight = consumed_weight + c_weight
            monthlyProductWeights[int(reg_time.month) - 1] = monthlyProductWeights[int(reg_time.month) - 1] + w
            monthlyUnconsumedWeights[int(reg_time.month) - 1] = monthlyUnconsumedWeights[int(reg_time.month) - 1] + float(w - c_weight)

    unconsumed_weight = float("{0:.2f}".format(weight - consumed_weight))
    consumed_weight = float("{0:.2f}".format(consumed_weight))

    weight = float("{0:.2f}".format(weight))
    ambassadors = Ambassador.objects.all()



    cityList = []
    # users = Member.objects.filter(country='India').order_by('-id')
    users = Member.objects.all().order_by('-id')
    for user in users:
        products = Product.objects.filter(member_id=user.pk)
        user.products = str(products.count())
        city = user.city
        if not city in cityList:
            cityList.append(city)

    cities = []
    prodcount = []
    for city in cityList:
        members = Member.objects.filter(city=city)
        products = Product.objects.filter(city=city)

        prodcount.append(products.count())

        cityInfo = {
            'city':city,
            'users':members.count(),
            'products':products.count()
        }

        cities.append(cityInfo)

    topcities = []

    if len(prodcount) > 0:

        prodcount.sort(reverse=True)

        i = 0
        for pcnt in prodcount:
            for city in cities:
                if pcnt == city['products'] and not city in topcities:
                    if city['products'] > 0:
                        topcities.append(city)
                    i = i + 1
                    if i == 10: break

    topusers = []
    prodcount = []
    for user in users:
        products = Product.objects.filter(member_id=user.pk)
        user.products = str(products.count())
        prodcount.append(products.count())
    if len(prodcount) > 0:
        prodcount.sort(reverse=True)
        i = 0
        for pcnt in prodcount:
            for user in users:
                if pcnt == int(user.products) and not user in topusers:
                    if int(user.products) > 0:
                        topusers.append(user)
                    i = i + 1
                    if i == 10: break

    # return HttpResponse(prodcount)


    context = {
        'monthlyusercount':monthlyUserCounts,
        'monthlysharedusercount':monthlySharedUserCounts,
        'monthlyproductcount':monthlyProductCounts,
        'allusercount':users.count(),
        'sharedusercount':sharedusercount,
        'allproductcount':allproducts.count(),
        'productweight': weight,
        'ambassadorcount':ambassadors.count(),
        'unconsumedweight':unconsumed_weight,
        'consumedweight':consumed_weight,
        'monthlyProductWeights':monthlyProductWeights,
        'monthlyUnconsumedWeights':monthlyUnconsumedWeights,
        'users':users,
        'cities':cities,
        'topcities':topcities,
        'topusers':topusers
    }

    return render(request, 'sharemyfood/analysis.html', context)




def emm(request):

    mms = Mm.objects.filter(id=1)
    mm = None
    if mms.count() == 0:
        mm = Mm()
        mm.value = '0'
        mm.save()
    else:
        mm = mms[0]
    mm.value = '1'
    mm.save()

    return HttpResponse('<h3>Locked sharemyfood!</h3>')


def dmm(request):

    mms = Mm.objects.filter(id=1)
    mm = None
    if mms.count() == 0:
        mm = Mm()
        mm.value = '0'
        mm.save()
    else:
        mm = mms[0]
    mm.value = '0'
    mm.save()

    return HttpResponse('<h3>Unlocked sharemyfood!</h3>')



























































